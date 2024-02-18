from jiwer import wer
import os
import playsound
import speech_recognition as sr
from gtts import gTTS
import pandas as pd
from openpyxl import Workbook, load_workbook
import matplotlib.pyplot as plt


def phat_am(text):
    print("Trợ Lý ảo:  ", text)
    tts = gTTS(text=text, lang="vi", slow=False)
    tts.save("sound.mp3")
    playsound.playsound("sound.mp3", True)
    os.remove("sound.mp3")


def lay_am_thanh(duong_dan_thu_muc):
    recognizer = sr.Recognizer()
    thu_muc_am_thanh = os.path.join(duong_dan_thu_muc, "recordedsound")

    os.makedirs(thu_muc_am_thanh, exist_ok=True)

    so = 1
    while os.path.exists(os.path.join(thu_muc_am_thanh, f"ghi_am_{so}.wav")):
        so += 1

    ten_tap_tin_am_thanh = f"ghi_am_{so}.wav"
    duong_dan_tap_tin_am_thanh = os.path.join(
        thu_muc_am_thanh, ten_tap_tin_am_thanh)

    with sr.Microphone() as nguon:
        print("Trợ Lý Ảo: Đang nghe! -- __ --")
        am_thanh = recognizer.listen(nguon, timeout=None)

        try:
            print("Trợ Lý Ảo: ...")
            van_ban = recognizer.recognize_google(am_thanh, language="vi-VN")
            print("Tôi: ", van_ban)

            with open(duong_dan_tap_tin_am_thanh, "wb") as tap_tin_am_thanh:
                tap_tin_am_thanh.write(am_thanh.get_wav_data())

            return van_ban.lower().replace(",", ""), ten_tap_tin_am_thanh, so
        except Exception as ex:
            print("Trợ Lý Ảo: Lỗi Rồi! ...")
            return 0, 0


def dung():
    phat_am("Hẹn gặp lại sau nha ! ... ")


def kiem_tra_ton_tai(duong_dan_thu_muc):
    return os.path.exists(duong_dan_thu_muc)


def tao_thu_muc_moi(duong_dan, ten_thu_muc):
    os.makedirs(os.path.join(duong_dan, ten_thu_muc))


def danh_gia_cau_hoi(ten_thu_muc, so_dong_excel):
    duong_dan_excel_output = os.path.join(ten_thu_muc, "recordings.xlsx")

    if os.path.exists(duong_dan_excel_output):
        workbook = load_workbook(filename=duong_dan_excel_output)
        sheet = workbook.active
    else:
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["Số thứ tự", "File Name", "Original Text",
                     "Recognized Text", "Corrected Rate"])

    duong_dan_excel = r"C:\Users\buildpc.shop\Desktop\speech2\1000sentences.xlsx"
    df = pd.read_excel(duong_dan_excel, header=None, usecols=[4])

    try:
        cau_hoi_goc = df.iloc[so_dong_excel - 1, 0].lower().replace(
            ",", "").replace(".", "").replace(" ", "   ")
    except IndexError:
        phat_am("Số thứ tự không hợp lệ trong Excel.")
        return

    print(f"Câu số {so_dong_excel - 2}: {cau_hoi_goc.capitalize()}")

    van_ban_noi, so_am_thanh, so_thu_tu = lay_am_thanh(ten_thu_muc)
    if van_ban_noi:
        van_ban_noi = van_ban_noi.lower().replace(",", "").replace(".", "")

        gia_tri_wer = wer(cau_hoi_goc, van_ban_noi) * 100
        sheet.append([so_thu_tu, so_am_thanh, cau_hoi_goc,
                     van_ban_noi, 100 - gia_tri_wer])
        workbook.save(filename=duong_dan_excel_output)

        print(f"Độ chính xác của bạn là: {100 - gia_tri_wer:.2f}%")
    else:
        phat_am("Không nhận diện được âm thanh. Thử lại sau.")


ten_nguoi_dung = input("Nhập tên của bạn: ")
duong_dan_goc = r"C:\Users\buildpc.shop\Desktop\speech2"
duong_dan_thu_muc = os.path.join(duong_dan_goc, ten_nguoi_dung)

if not kiem_tra_ton_tai(duong_dan_thu_muc):
    tao_thu_muc_moi(duong_dan_goc, ten_nguoi_dung)

while True:
    so_dong_excel = input(
        "Nhập số thứ tự dòng cần kiểm tra trong Excel (nhập 'dung' để kết thúc): ")
    if so_dong_excel.lower() == "dung":
        dung()
        break
    else:
        so_dong_excel = int(so_dong_excel) + 2
        danh_gia_cau_hoi(duong_dan_thu_muc, so_dong_excel)

duong_dan_excel_output = os.path.join(duong_dan_thu_muc, "recordings.xlsx")
if os.path.exists(duong_dan_excel_output):
    df = pd.read_excel(duong_dan_excel_output)

    # Plotting the Corrected Rate against File Name
    plt.figure(figsize=(10, 6))
    bar = plt.bar(df['File Name'], df['Corrected Rate'], color='orange')
    plt.xlabel('File Name')
    plt.ylabel('Corrected Rate')
    plt.title('Corrected Rate vs File Name')
    plt.xticks(rotation=0)
    plt.tight_layout()
    plt.grid(True)
    for bar, rate in zip(bar, df['Corrected Rate']):
        plt.text(bar.get_x() + bar.get_width() / 2, bar.get_height(), f'{rate:.2f}%',
                 ha='center', va='bottom', fontsize=8, color='black')
    plt.show()
else:
    print("Excel file not found.")
